"""
Export service for time entries (Timeliste) to PDF and Excel formats.
"""

from datetime import datetime, timedelta
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class TimeEntriesExportService:
    """Service for exporting time entries to various formats."""

    @staticmethod
    def calculate_date_range(period):
        """
        Calculate start and end dates for various time periods.

        Args:
            period (str): One of: 'this_week', 'last_week', 'this_month', 'last_month', 'all_time'

        Returns:
            tuple: (start_date, end_date, period_description) or (None, None, 'All Time') for all_time
        """
        now = timezone.now()
        today = now.date()

        if period == "this_week":
            # Start of current week (Monday)
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)  # Sunday
            description = (
                f"This Week ({start.strftime('%b %d')} - {end.strftime('%b %d, %Y')})"
            )
            return start, end, description

        elif period == "last_week":
            # Start of last week (Monday)
            start = today - timedelta(days=today.weekday() + 7)
            end = start + timedelta(days=6)  # Sunday
            description = (
                f"Last Week ({start.strftime('%b %d')} - {end.strftime('%b %d, %Y')})"
            )
            return start, end, description

        elif period == "this_month":
            # First day of current month
            start = today.replace(day=1)
            # Last day of current month
            if today.month == 12:
                end = today.replace(month=12, day=31)
            else:
                end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
            description = f"This Month ({start.strftime('%B %Y')})"
            return start, end, description

        elif period == "last_month":
            # First day of last month
            first_of_current_month = today.replace(day=1)
            last_month_end = first_of_current_month - timedelta(days=1)
            start = last_month_end.replace(day=1)
            end = last_month_end
            description = f"Last Month ({start.strftime('%B %Y')})"
            return start, end, description

        elif period == "all_time":
            return None, None, "All Time"

        else:
            raise ValueError(
                f"Invalid period: {period}. Must be one of: this_week, last_week, this_month, last_month, all_time"
            )

    @staticmethod
    def generate_excel(queryset, period, user):
        """
        Generate an Excel workbook from time entries queryset.

        Args:
            queryset: Django queryset of Timeliste objects
            period: Period string (for filename and header)
            user: User object for the report

        Returns:
            BytesIO: Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Entries"

        # Calculate date range info
        start_date, end_date, period_desc = (
            TimeEntriesExportService.calculate_date_range(period)
        )

        # Header styling
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Time Entries Report - {user.username}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Period info
        ws.merge_cells("A2:E2")
        ws["A2"] = f"Period: {period_desc}"
        ws["A2"].font = Font(size=11)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Generated date
        ws.merge_cells("A3:E3")
        ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A3"].font = Font(size=10, italic=True)
        ws["A3"].alignment = Alignment(horizontal="center")

        # Empty row
        ws.append([])

        # Table headers
        headers = ["Date", "Job", "Description", "Hours", "Created At"]
        ws.append(headers)

        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data rows
        total_hours = 0
        for entry in queryset.select_related("jobb", "user").order_by(
            "dato", "created_at"
        ):
            # Convert minutes to hours (timer field stores minutes)
            hours = entry.timer / 60 if entry.timer else 0
            total_hours += hours

            row = [
                entry.dato.strftime("%Y-%m-%d") if entry.dato else "N/A",
                entry.jobb.tittel if entry.jobb else "N/A",
                entry.beskrivelse or "",
                round(hours, 2),
                (
                    entry.created_at.strftime("%Y-%m-%d %H:%M")
                    if entry.created_at
                    else "N/A"
                ),
            ]
            ws.append(row)

        # Summary row
        summary_row = ws.max_row + 2
        ws.merge_cells(f"A{summary_row}:C{summary_row}")
        ws[f"A{summary_row}"] = "TOTAL HOURS:"
        ws[f"A{summary_row}"].font = Font(bold=True, size=12)
        ws[f"A{summary_row}"].alignment = Alignment(horizontal="right")

        ws[f"D{summary_row}"] = round(total_hours, 2)
        ws[f"D{summary_row}"].font = Font(bold=True, size=12)
        ws[f"D{summary_row}"].fill = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
        )

        # Entry count
        count_row = summary_row + 1
        ws.merge_cells(f"A{count_row}:C{count_row}")
        ws[f"A{count_row}"] = "TOTAL ENTRIES:"
        ws[f"A{count_row}"].font = Font(bold=True)
        ws[f"A{count_row}"].alignment = Alignment(horizontal="right")

        ws[f"D{count_row}"] = queryset.count()
        ws[f"D{count_row}"].font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 18

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_pdf(queryset, period, user):
        """
        Generate a PDF report from time entries queryset.

        Args:
            queryset: Django queryset of Timeliste objects
            period: Period string (for filename and header)
            user: User object for the report

        Returns:
            BytesIO: PDF file as bytes
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch
        )

        # Container for the 'Flowable' objects
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#366092"),
            spaceAfter=12,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle", parent=styles["Normal"], fontSize=12, spaceAfter=6
        )
        info_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20,
        )

        # Calculate date range info
        start_date, end_date, period_desc = (
            TimeEntriesExportService.calculate_date_range(period)
        )

        # Title
        title = Paragraph(f"Time Entries Report", title_style)
        elements.append(title)

        # User info
        user_info = Paragraph(
            f"<b>User:</b> {user.username} ({user.email})", subtitle_style
        )
        elements.append(user_info)

        # Period info
        period_info = Paragraph(f"<b>Period:</b> {period_desc}", subtitle_style)
        elements.append(period_info)

        # Generated date
        gen_date = Paragraph(
            f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", info_style
        )
        elements.append(gen_date)

        # Summary statistics
        total_hours = 0
        entries_by_date = {}

        for entry in queryset.select_related("jobb", "user").order_by(
            "dato", "created_at"
        ):
            # Convert minutes to hours
            hours = entry.timer / 60 if entry.timer else 0
            total_hours += hours

            # Group by date
            date_str = entry.dato.strftime("%Y-%m-%d") if entry.dato else "No Date"
            if date_str not in entries_by_date:
                entries_by_date[date_str] = []
            entries_by_date[date_str].append(entry)

        # Summary box
        summary_data = [
            ["Summary", ""],
            ["Total Hours:", f"{round(total_hours, 2)}h"],
            ["Total Entries:", str(queryset.count())],
            ["Unique Days:", str(len(entries_by_date))],
        ]

        summary_table = Table(summary_data, colWidths=[2.5 * inch, 1.5 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ]
            )
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Detailed entries table
        if queryset.exists():
            # Table header
            table_data = [["Date", "Job", "Description", "Hours"]]

            # Add entries
            for entry in queryset.select_related("jobb").order_by("dato", "created_at"):
                hours = entry.timer / 60 if entry.timer else 0
                row = [
                    entry.dato.strftime("%Y-%m-%d") if entry.dato else "N/A",
                    entry.jobb.tittel[:30] if entry.jobb else "N/A",
                    entry.beskrivelse[:50] if entry.beskrivelse else "",
                    f"{round(hours, 2)}h",
                ]
                table_data.append(row)

            # Create table
            col_widths = [1.2 * inch, 2 * inch, 2.8 * inch, 0.8 * inch]
            entries_table = Table(table_data, colWidths=col_widths, repeatRows=1)

            entries_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (3, 0), (3, -1), "RIGHT"),  # Right-align hours column
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.lightgrey],
                        ),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            elements.append(entries_table)
        else:
            no_data = Paragraph(
                "<i>No time entries found for this period.</i>", styles["Normal"]
            )
            elements.append(no_data)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
